from __future__ import annotations

import json
import re
import tempfile
import subprocess
import sys
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "EverBatt 2023.xlsm"
OUT_DIR = ROOT / "docs" / "analysis"

NS_MAIN = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
NS_REL = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}


def cell_ref(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"


def first_nonempty(values: list[str], limit: int = 8) -> list[str]:
    seen = []
    for value in values:
        if value and value not in seen:
            seen.append(value)
        if len(seen) >= limit:
            break
    return seen


def extract_workbook_xml() -> tuple[list[dict], dict[str, str], list[dict]]:
    with zipfile.ZipFile(WORKBOOK) as archive:
        workbook_xml = ET.fromstring(archive.read("xl/workbook.xml"))
        rels_xml = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))

    rel_targets = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels_xml.findall("r:Relationship", NS_REL)
    }

    sheets = []
    for sheet in workbook_xml.findall("m:sheets/m:sheet", NS_MAIN):
        rel_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        sheets.append(
            {
                "name": sheet.attrib["name"],
                "sheet_id": sheet.attrib.get("sheetId"),
                "state": sheet.attrib.get("state", "visible"),
                "target": f"xl/{rel_targets.get(rel_id, '')}",
            }
        )

    defined_names = []
    for defined in workbook_xml.findall("m:definedNames/m:definedName", NS_MAIN):
        defined_names.append(
            {
                "name": defined.attrib.get("name"),
                "local_sheet_id": defined.attrib.get("localSheetId"),
                "hidden": defined.attrib.get("hidden"),
                "text": defined.text,
            }
        )

    return sheets, rel_targets, defined_names


def workbook_for_openpyxl() -> Path:
    """Create a temporary copy with stylesheet values normalized for openpyxl."""
    with zipfile.ZipFile(WORKBOOK) as archive:
        def clamp_family(match: re.Match[bytes]) -> bytes:
            value = int(match.group(1))
            if value <= 14:
                return match.group(0)
            return b'<family val="14"'

        tmp = Path(tempfile.gettempdir()) / "everbatt_openpyxl_readable.xlsm"
        with zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as out:
            for info in archive.infolist():
                data = archive.read(info.filename)
                if info.filename.endswith(".xml"):
                    data = re.sub(rb'<family val="([0-9]+)"', clamp_family, data)
                out.writestr(info, data)

    return tmp


def summarize_sheet(ws, formulas_ws) -> dict:
    nonempty = []
    formulas = []
    input_candidates = []
    text_values = []

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            ref = cell.coordinate
            nonempty.append(ref)
            if isinstance(cell.value, str):
                text_values.append(cell.value.strip())

            formula_cell = formulas_ws[ref]
            if isinstance(formula_cell.value, str) and formula_cell.value.startswith("="):
                formula = formula_cell.value
                formulas.append({"cell": ref, "formula": formula})
            elif not isinstance(cell.value, str) or not cell.value.startswith("="):
                fill = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
                input_candidates.append(
                    {
                        "cell": ref,
                        "value": cell.value,
                        "number_format": cell.number_format,
                        "fill": fill,
                    }
                )

    formula_heads = Counter()
    formula_refs = Counter()
    for item in formulas:
        formula = item["formula"]
        head = re.match(r"=([A-Z][A-Z0-9_.]*)\\b", formula, flags=re.I)
        if head:
            formula_heads[head.group(1).upper()] += 1
        for ref in re.findall(r"(?:'[^']+'|[A-Za-z0-9_ ]+)!\$?[A-Z]{1,3}\$?\d+", formula):
            formula_refs[ref.split("!")[0].strip("'")] += 1

    return {
        "title_values": first_nonempty(text_values, 12),
        "dimensions": ws.calculate_dimension(),
        "nonempty_cells": len(nonempty),
        "formula_cells": len(formulas),
        "formula_functions_top": formula_heads.most_common(20),
        "formula_sheet_refs_top": formula_refs.most_common(20),
        "sample_formulas": formulas[:80],
        "sample_input_candidates": input_candidates[:120],
        "tables": [table.name for table in ws.tables.values()],
        "merged_ranges": [str(rng) for rng in list(ws.merged_cells.ranges)[:80]],
        "data_validations": len(ws.data_validations.dataValidation),
    }


def extract_zip_parts() -> dict:
    counts = Counter()
    controls = []
    drawings = []
    comments = {}
    connections = None

    with zipfile.ZipFile(WORKBOOK) as archive:
        for name in archive.namelist():
            top = name.split("/")[1] if "/" in name else name
            counts[top] += 1
            if name.endswith(".xml"):
                raw = archive.read(name)
                if b"control" in raw.lower() or b"macro" in raw.lower():
                    controls.append(name)
                if name.startswith("xl/drawings/drawing"):
                    drawings.append(name)
                if name.startswith("xl/comments") and name.endswith(".xml"):
                    try:
                        root = ET.fromstring(raw)
                        texts = []
                        for t in root.findall(".//m:t", NS_MAIN):
                            if t.text:
                                texts.append(t.text)
                        comments[name] = first_nonempty([x.strip() for x in texts], 20)
                    except ET.ParseError:
                        pass
                if name == "xl/connections.xml":
                    connections = raw.decode("utf-8", errors="replace")

    return {
        "zip_part_counts": counts,
        "control_or_macro_xml_parts": controls,
        "drawing_parts": drawings,
        "comments": comments,
        "connections_excerpt": connections[:4000] if connections else None,
    }


def extract_vba() -> dict:
    out_file = OUT_DIR / "vba_olevba.txt"
    cmd = [sys.executable, "-m", "oletools.olevba", str(WORKBOOK)]
    proc = subprocess.run(cmd, cwd=ROOT, text=True, capture_output=True, check=False)
    text = proc.stdout + proc.stderr
    out_file.write_text(text, encoding="utf-8", errors="replace")

    modules = []
    current = None
    for line in text.splitlines():
        if line.startswith("VBA MACRO "):
            if current:
                modules.append(current)
            current = {"header": line, "lines": []}
        elif current is not None:
            current["lines"].append(line)
    if current:
        modules.append(current)

    module_summaries = []
    for module in modules:
        code = "\n".join(module["lines"])
        procedures = re.findall(r"^\s*(?:Public |Private )?(Sub|Function)\s+([A-Za-z_][A-Za-z0-9_]*)", code, re.M)
        calls = Counter(
            re.findall(
                r"\b(Application\.[A-Za-z_]+|Worksheets?\(|Range\(|Cells\(|Calculate|GoalSeek|Solver[A-Za-z]*)",
                code,
            )
        )
        module_summaries.append(
            {
                "header": module["header"],
                "line_count": len(module["lines"]),
                "procedures": [{"type": kind, "name": name} for kind, name in procedures],
                "excel_api_calls_top": calls.most_common(20),
            }
        )

    return {
        "olevba_returncode": proc.returncode,
        "olevba_output": str(out_file.relative_to(ROOT)),
        "modules": module_summaries,
    }


def build_dependency_edges(wb_formulas) -> dict:
    edges = defaultdict(Counter)
    by_sheet_formula_count = Counter()
    for ws in wb_formulas.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not (isinstance(value, str) and value.startswith("=")):
                    continue
                by_sheet_formula_count[ws.title] += 1
                for ref_sheet in re.findall(r"(?:'([^']+)'|([A-Za-z0-9_ ]+))!\$?[A-Z]{1,3}\$?\d+", value):
                    sheet = ref_sheet[0] or ref_sheet[1]
                    if sheet != ws.title:
                        edges[ws.title][sheet] += 1
    return {
        "formula_count_by_sheet": by_sheet_formula_count,
        "cross_sheet_edges": {
            sheet: refs.most_common()
            for sheet, refs in sorted(edges.items())
        },
    }


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    sheets, _rels, defined_names = extract_workbook_xml()
    readable_workbook = workbook_for_openpyxl()
    wb_values = load_workbook(readable_workbook, data_only=True, keep_vba=True)
    wb_formulas = load_workbook(readable_workbook, data_only=False, keep_vba=True)

    sheet_summaries = {}
    for sheet in sheets:
        name = sheet["name"]
        sheet_summaries[name] = summarize_sheet(wb_values[name], wb_formulas[name])

    report = {
        "workbook": str(WORKBOOK.name),
        "sheets": sheets,
        "defined_names_count": len(defined_names),
        "defined_names_sample": defined_names[:200],
        "sheet_summaries": sheet_summaries,
        "dependencies": build_dependency_edges(wb_formulas),
        "package": extract_zip_parts(),
        "vba": extract_vba(),
    }

    report_path = OUT_DIR / "everbatt_analysis.json"
    report_path.write_text(json.dumps(report, indent=2, ensure_ascii=False, default=str), encoding="utf-8")

    print(f"Wrote {report_path.relative_to(ROOT)}")
    print("Sheets:")
    for sheet in sheets:
        summary = sheet_summaries[sheet["name"]]
        print(
            f"- {sheet['name']}: {summary['dimensions']}, "
            f"{summary['nonempty_cells']} non-empty, {summary['formula_cells']} formulas"
        )
    print(f"VBA modules: {len(report['vba']['modules'])}; raw output: {report['vba']['olevba_output']}")


if __name__ == "__main__":
    main()
