from __future__ import annotations

import argparse
import json
import math
import subprocess
import sys
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

from recycle_cost.model import FeedstockInput, Scenario, TransportDistances, default_scenario  # noqa: E402
from recycle_cost.reporting import python_ported_output_summary_table  # noqa: E402
from recycle_cost.schemas import CommonColumns  # noqa: E402
from recycle_cost.workbook import WORKBOOK_PATH, make_openpyxl_readable_copy  # noqa: E402


NS_MAIN = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
NS_REL = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

PROCESS_COL = {"Pyro": "J", "Hydro": "K", "Direct": "L", "Custom": "M"}
PROCESS_LABEL = {
    "Pyro": "Pyrometallurgical",
    "Hydro": "Hydrometallurgical",
    "Direct": "Direct",
    "Custom": "Custom",
}

SELECTED_RECYCLING_METRICS = {
    "Recycling cost": 37,
    "Recycling total energy": 39,
    "Recycling water": 44,
    "Recycling revenue": 59,
    "Recycling GHGs": 58,
}

OUTPUT_SUMMARY_CELLS = [
    ("Cell manufacturing cost", "Virgin", "C11"),
    ("Cell manufacturing cost", "Pyro", "D11"),
    ("Cell manufacturing cost", "Hydro", "E11"),
    ("Cell manufacturing cost", "Direct", "F11"),
    ("Cell manufacturing cost", "Custom", "G11"),
    ("Cell manufacturing total energy", "Virgin", "C13"),
    ("Cell manufacturing total energy", "Pyro", "D13"),
    ("Cell manufacturing total energy", "Hydro", "E13"),
    ("Cell manufacturing total energy", "Direct", "F13"),
    ("Cell manufacturing total energy", "Custom", "G13"),
    ("Cell manufacturing water", "Virgin", "C18"),
    ("Cell manufacturing water", "Pyro", "D18"),
    ("Cell manufacturing water", "Hydro", "E18"),
    ("Cell manufacturing water", "Direct", "F18"),
    ("Cell manufacturing water", "Custom", "G18"),
    ("Cell manufacturing GHGs", "Virgin", "C32"),
    ("Cell manufacturing GHGs", "Pyro", "D32"),
    ("Cell manufacturing GHGs", "Hydro", "E32"),
    ("Cell manufacturing GHGs", "Direct", "F32"),
    ("Cell manufacturing GHGs", "Custom", "G32"),
    ("Collection and transport cost", "Virgin", "C37"),
    ("Collection and transport total energy", "Virgin", "C39"),
]

for metric, row in SELECTED_RECYCLING_METRICS.items():
    for process, column in PROCESS_COL.items():
        OUTPUT_SUMMARY_CELLS.append((metric, process, f"{column}{row}"))


MATRIX_SCENARIOS = [
    {
        "name": "default_black_mass_hydro",
        "process": "Hydro",
        "battery_manufactured": "Pack",
        "manufacturing_chemistry": "NMC(622)",
        "manufacturing_location": "U.S.",
        "battery_collected": "Select Battery",
        "cathode_chemistry": "NMC(622)",
        "throughput": 50,
        "cathode_throughput": 0,
        "recycled_content": 0,
        "feedstocks": [("NMC(622)", "Black mass", 10000)],
    },
    {
        "name": "black_mass_pyro_nmc622",
        "process": "Pyro",
        "battery_manufactured": "Pack",
        "manufacturing_chemistry": "NMC(622)",
        "manufacturing_location": "U.S.",
        "battery_collected": "Select Battery",
        "cathode_chemistry": "NMC(622)",
        "throughput": 50,
        "cathode_throughput": 0,
        "recycled_content": 0,
        "feedstocks": [("NMC(622)", "Black mass", 10000)],
    },
    {
        "name": "black_mass_direct_nmc622",
        "process": "Direct",
        "battery_manufactured": "Pack",
        "manufacturing_chemistry": "NMC(622)",
        "manufacturing_location": "U.S.",
        "battery_collected": "Select Battery",
        "cathode_chemistry": "NMC(622)",
        "throughput": 50,
        "cathode_throughput": 0,
        "recycled_content": 0,
        "feedstocks": [("NMC(622)", "Black mass", 10000)],
    },
    {
        "name": "black_mass_direct_lfp",
        "process": "Direct",
        "battery_manufactured": "Pack",
        "manufacturing_chemistry": "LFP",
        "manufacturing_location": "U.S.",
        "battery_collected": "Select Battery",
        "cathode_chemistry": "LFP",
        "throughput": 50,
        "cathode_throughput": 0,
        "recycled_content": 0,
        "feedstocks": [("LFP", "Black mass", 9000)],
    },
    {
        "name": "pack_pyro_nmc622",
        "process": "Pyro",
        "battery_manufactured": "Pack",
        "manufacturing_chemistry": "NMC(622)",
        "manufacturing_location": "U.S.",
        "battery_collected": "Module",
        "cathode_chemistry": "NMC(622)",
        "throughput": 50,
        "cathode_throughput": 10,
        "recycled_content": 0.2,
        "feedstocks": [("NMC(622)", "End-of-life battery: pack", 10000)],
    },
    {
        "name": "pack_pyro_lfp",
        "process": "Pyro",
        "battery_manufactured": "Pack",
        "manufacturing_chemistry": "LFP",
        "manufacturing_location": "U.S.",
        "battery_collected": "Module",
        "cathode_chemistry": "LFP",
        "throughput": 50,
        "cathode_throughput": 10,
        "recycled_content": 0.2,
        "feedstocks": [("LFP", "End-of-life battery: pack", 10000)],
    },
    {
        "name": "pack_hydro_nmc622",
        "process": "Hydro",
        "battery_manufactured": "Pack",
        "manufacturing_chemistry": "NMC(622)",
        "manufacturing_location": "U.S.",
        "battery_collected": "Module",
        "cathode_chemistry": "NMC(622)",
        "throughput": 50,
        "cathode_throughput": 10,
        "recycled_content": 0.2,
        "feedstocks": [("NMC(622)", "End-of-life battery: pack", 10000)],
    },
    {
        "name": "pack_hydro_nmc811",
        "process": "Hydro",
        "battery_manufactured": "Pack",
        "manufacturing_chemistry": "NMC(811)",
        "manufacturing_location": "U.S.",
        "battery_collected": "Module",
        "cathode_chemistry": "NMC(811)",
        "throughput": 50,
        "cathode_throughput": 10,
        "recycled_content": 0.2,
        "feedstocks": [("NMC(811)", "End-of-life battery: pack", 10000)],
    },
    {
        "name": "pack_direct_nmc622",
        "process": "Direct",
        "battery_manufactured": "Pack",
        "manufacturing_chemistry": "NMC(622)",
        "manufacturing_location": "U.S.",
        "battery_collected": "Module",
        "cathode_chemistry": "NMC(622)",
        "throughput": 50,
        "cathode_throughput": 10,
        "recycled_content": 0.2,
        "feedstocks": [("NMC(622)", "End-of-life battery: pack", 10000)],
    },
    {
        "name": "pack_direct_nca",
        "process": "Direct",
        "battery_manufactured": "Pack",
        "manufacturing_chemistry": "NCA",
        "manufacturing_location": "U.S.",
        "battery_collected": "Module",
        "cathode_chemistry": "NCA",
        "throughput": 50,
        "cathode_throughput": 9,
        "recycled_content": 0.2,
        "feedstocks": [("NCA", "End-of-life battery: pack", 9000)],
    },
    {
        "name": "module_hydro_nmc622",
        "process": "Hydro",
        "battery_manufactured": "Module",
        "manufacturing_chemistry": "NMC(622)",
        "manufacturing_location": "U.S.",
        "battery_collected": "Cell",
        "cathode_chemistry": "NMC(622)",
        "throughput": 50,
        "cathode_throughput": 8,
        "recycled_content": 0.15,
        "feedstocks": [("NMC(622)", "End-of-life battery: module", 8000)],
    },
    {
        "name": "cell_pyro_nmc622",
        "process": "Pyro",
        "battery_manufactured": "Cell",
        "manufacturing_chemistry": "NMC(622)",
        "manufacturing_location": "U.S.",
        "battery_collected": "Cell",
        "cathode_chemistry": "NMC(622)",
        "throughput": 50,
        "cathode_throughput": 6,
        "recycled_content": 0.1,
        "feedstocks": [("NMC(622)", "End-of-life battery: cell", 7000)],
    },
    {
        "name": "scrap_rejected_hydro_nmc622",
        "process": "Hydro",
        "battery_manufactured": "Cell",
        "manufacturing_chemistry": "NMC(622)",
        "manufacturing_location": "U.S.",
        "battery_collected": "Cell",
        "cathode_chemistry": "NMC(622)",
        "throughput": 50,
        "cathode_throughput": 6,
        "recycled_content": 0.2,
        "feedstocks": [("NMC(622)", "Manufacturing scrap: rejected cells", 6000)],
    },
    {
        "name": "scrap_electrode_direct_nmc622",
        "process": "Direct",
        "battery_manufactured": "Cell",
        "manufacturing_chemistry": "NMC(622)",
        "manufacturing_location": "U.S.",
        "battery_collected": "Cell",
        "cathode_chemistry": "NMC(622)",
        "throughput": 50,
        "cathode_throughput": 10,
        "recycled_content": 0.2,
        "feedstocks": [("NMC(622)", "Manufacturing scrap: electrode", 5000)],
    },
    {
        "name": "mixed_pack_scrap_hydro",
        "process": "Hydro",
        "battery_manufactured": "Pack",
        "manufacturing_chemistry": "NMC(622)",
        "manufacturing_location": "U.S.",
        "battery_collected": "Module",
        "cathode_chemistry": "NMC(622)",
        "throughput": 50,
        "cathode_throughput": 10,
        "recycled_content": 0.2,
        "feedstocks": [
            ("NMC(622)", "End-of-life battery: pack", 7000),
            ("NMC(811)", "Manufacturing scrap: electrode", 3000),
        ],
    },
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Compare Python output summaries with LibreOffice-recalculated EverBatt matrix workbooks."
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=Path("/tmp/everbatt_matrix_inputs"),
        help="Directory containing scenario input .xlsm workbooks.",
    )
    parser.add_argument(
        "--recalc-dir",
        type=Path,
        default=Path("/tmp/everbatt_matrix_recalc"),
        help="Directory containing recalculated .xlsx workbooks, or output directory when --recalculate is set.",
    )
    parser.add_argument(
        "--out-dir",
        type=Path,
        default=Path("/tmp/recycle_cost_excel_compare"),
        help="Directory for comparison CSV files and summary.json.",
    )
    parser.add_argument(
        "--recalculate",
        action="store_true",
        help="Run LibreOffice headless to regenerate .xlsx files from --input-dir before comparing.",
    )
    parser.add_argument(
        "--generate-inputs",
        action="store_true",
        help="Generate the built-in 15-scenario .xlsm matrix into --input-dir before comparing.",
    )
    parser.add_argument(
        "--workbook-template",
        type=Path,
        default=WORKBOOK_PATH,
        help="EverBatt .xlsm workbook used as the source template with --generate-inputs.",
    )
    parser.add_argument("--soffice", default="soffice", help="LibreOffice executable used with --recalculate.")
    parser.add_argument(
        "--libreoffice-profile",
        type=Path,
        default=Path("/tmp/lo-profile-recycle-cost-compare"),
        help="LibreOffice user profile directory used with --recalculate.",
    )
    parser.add_argument(
        "--selected-tolerance",
        type=float,
        default=1e-9,
        help="Tolerance used for selected-route recycling parity checks.",
    )
    parser.add_argument(
        "--full-tolerance",
        type=float,
        default=1e-6,
        help="Tolerance used for the broader Output summary diagnostic.",
    )
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Exit with status 1 when selected-route recycling exceeds --selected-tolerance.",
    )
    return parser.parse_args()


def clean_number(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped or stripped.startswith("#"):
            return None
        try:
            value = float(stripped)
        except ValueError:
            return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if math.isnan(number) or math.isinf(number):
        return None
    return number


def clean_text(value) -> str:
    return "" if value is None else str(value).strip()


def override_or_default(ws, row: int, default: float | None) -> float | None:
    override = clean_number(ws[f"G{row}"].value)
    fallback = clean_number(ws[f"F{row}"].value)
    if override is not None:
        return override
    if fallback is not None:
        return fallback
    return default


def run_libreoffice_recalc(args: argparse.Namespace) -> None:
    input_files = sorted(args.input_dir.glob("*.xlsm"))
    if not input_files:
        raise FileNotFoundError(f"No .xlsm files found in {args.input_dir}")

    args.recalc_dir.mkdir(parents=True, exist_ok=True)
    args.libreoffice_profile.mkdir(parents=True, exist_ok=True)
    cmd = [
        args.soffice,
        "--headless",
        "--nologo",
        "--nodefault",
        "--nofirststartwizard",
        "--nolockcheck",
        f"-env:UserInstallation=file://{args.libreoffice_profile}",
        "--convert-to",
        "xlsx",
        "--outdir",
        str(args.recalc_dir),
        *[str(path) for path in input_files],
    ]
    subprocess.run(cmd, cwd=ROOT, check=True)


def clear_feedstock_rows(ws) -> None:
    for row in range(28, 33):
        ws[f"C{row}"] = "Select Chemistry"
        ws[f"D{row}"] = "Select Type"
        ws[f"F{row}"] = 0


def write_matrix_scenario(workbook, scenario: dict[str, object]) -> None:
    input_ws = workbook["Input"]
    output_ws = workbook["Output"]

    input_ws["E6"] = scenario["battery_manufactured"]
    input_ws["G8"] = scenario["throughput"]
    input_ws["E9"] = scenario["manufacturing_chemistry"]
    input_ws["E10"] = scenario["manufacturing_location"]
    input_ws["E15"] = scenario["battery_collected"]
    input_ws["G17"] = 100
    input_ws["G18"] = 500
    input_ws["G19"] = 1000
    input_ws["G20"] = 1000
    input_ws["G21"] = 1000
    input_ws["G22"] = 1000
    input_ws["E38"] = scenario["cathode_chemistry"]
    input_ws["G48"] = scenario["cathode_throughput"]
    input_ws["G49"] = scenario["recycled_content"]

    clear_feedstock_rows(input_ws)
    for offset, feedstock in enumerate(scenario["feedstocks"]):
        row = 28 + offset
        chemistry, feedstock_type, tonnes = feedstock
        input_ws[f"C{row}"] = chemistry
        input_ws[f"D{row}"] = feedstock_type
        input_ws[f"F{row}"] = tonnes

    output_ws["J61"] = scenario["process"]
    output_ws["K61"] = None
    output_ws["L61"] = None
    output_ws["M61"] = None


def generate_matrix_inputs(args: argparse.Namespace) -> None:
    args.input_dir.mkdir(parents=True, exist_ok=True)
    readable_template = make_openpyxl_readable_copy(args.workbook_template)

    for scenario in MATRIX_SCENARIOS:
        workbook = load_workbook(readable_template, data_only=False, keep_vba=True)
        write_matrix_scenario(workbook, scenario)
        out_path = args.input_dir / f"{scenario['name']}.xlsm"
        workbook.save(out_path)


def shared_strings(archive: zipfile.ZipFile) -> list[str]:
    try:
        root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    except KeyError:
        return []

    strings = []
    for item in root.findall("m:si", NS_MAIN):
        parts = [text.text or "" for text in item.findall(".//m:t", NS_MAIN)]
        strings.append("".join(parts))
    return strings


def sheet_target(archive: zipfile.ZipFile, sheet_name: str) -> str:
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rel_targets = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels.findall("r:Relationship", NS_REL)
    }

    for sheet in workbook.findall("m:sheets/m:sheet", NS_MAIN):
        if sheet.attrib.get("name") == sheet_name:
            rel_id = sheet.attrib[f"{{{OFFICE_REL_NS}}}id"]
            target = rel_targets[rel_id]
            return target if target.startswith("xl/") else f"xl/{target.lstrip('/')}"

    raise KeyError(sheet_name)


def read_sheet_cells(path: Path, sheet_name: str, addresses: set[str]) -> dict[str, object]:
    with zipfile.ZipFile(path) as archive:
        strings = shared_strings(archive)
        target = sheet_target(archive, sheet_name)
        root = ET.fromstring(archive.read(target))

    values = {address: None for address in addresses}
    for cell in root.findall(".//m:c", NS_MAIN):
        address = cell.attrib.get("r")
        if address not in values:
            continue

        cell_type = cell.attrib.get("t")
        if cell_type == "inlineStr":
            values[address] = "".join(text.text or "" for text in cell.findall(".//m:t", NS_MAIN))
            continue

        raw = cell.find("m:v", NS_MAIN)
        if raw is None:
            continue

        value: object = raw.text
        if cell_type == "s":
            value = strings[int(value)]
        elif cell_type == "b":
            value = bool(int(value))
        values[address] = value

    return values


def scenario_from_input_workbook(path: Path, process: str) -> Scenario:
    base = default_scenario()
    workbook = load_workbook(path, data_only=True, keep_vba=True)
    ws = workbook["Input"]

    feedstocks = []
    for row in range(28, 33):
        chemistry = clean_text(ws[f"C{row}"].value)
        feedstock_type = clean_text(ws[f"D{row}"].value)
        tonnes = clean_number(ws[f"F{row}"].value) or 0.0
        if (
            chemistry
            and feedstock_type
            and chemistry != "Select Chemistry"
            and feedstock_type != "Select Type"
            and tonnes > 0
        ):
            feedstocks.append(FeedstockInput(chemistry, feedstock_type, tonnes))

    if not feedstocks:
        feedstocks = list(base.feedstocks)

    first_feedstock = feedstocks[0]
    cathode_chemistry = clean_text(ws["E38"].value) or first_feedstock.chemistry or base.cathode_chemistry

    return Scenario(
        battery_manufactured=clean_text(ws["E6"].value) or base.battery_manufactured,
        throughput_gwh_per_year=override_or_default(ws, 8, base.throughput_gwh_per_year),
        manufacturing_chemistry=clean_text(ws["E9"].value)
        or first_feedstock.chemistry
        or base.manufacturing_chemistry,
        manufacturing_location=clean_text(ws["E10"].value) or base.manufacturing_location,
        battery_collected=clean_text(ws["E15"].value) or base.battery_collected,
        feedstock_chemistry=first_feedstock.chemistry,
        feedstock_type=first_feedstock.feedstock_type,
        feedstock_tonnes_per_year=first_feedstock.tonnes_per_year,
        recycling_process=PROCESS_LABEL[process],
        cathode_chemistry=cathode_chemistry,
        recycled_content=override_or_default(ws, 49, base.recycled_content),
        cathode_throughput_gwh_per_year=override_or_default(ws, 48, base.cathode_throughput_gwh_per_year),
        transport_distances=TransportDistances(
            collection_to_disassembly=override_or_default(
                ws, 17, base.transport_distances.collection_to_disassembly
            ),
            disassembly_to_preprocessor=override_or_default(
                ws, 18, base.transport_distances.disassembly_to_preprocessor
            ),
            preprocessor_to_cm_recovery=override_or_default(
                ws, 19, base.transport_distances.preprocessor_to_cm_recovery
            ),
            manufacturer_to_preprocessor_or_cm_recovery=override_or_default(
                ws, 20, base.transport_distances.manufacturer_to_preprocessor_or_cm_recovery
            ),
            recycler_to_cathode_producer=override_or_default(
                ws, 21, base.transport_distances.recycler_to_cathode_producer
            ),
            cathode_producer_to_manufacturer=override_or_default(
                ws, 22, base.transport_distances.cathode_producer_to_manufacturer
            ),
        ),
        feedstocks=tuple(feedstocks),
    )


def compare_workbooks(input_dir: Path, recalc_dir: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    selected_rows = []
    full_rows = []
    output_cells = {address for _, _, address in OUTPUT_SUMMARY_CELLS}
    output_cells.update({"J61", "K61", "L61", "M61"})

    recalc_files = sorted(recalc_dir.glob("*.xlsx"))
    if not recalc_files:
        raise FileNotFoundError(f"No .xlsx files found in {recalc_dir}")

    for recalc_path in recalc_files:
        scenario_name = recalc_path.stem
        input_path = input_dir / f"{scenario_name}.xlsm"
        if not input_path.exists():
            raise FileNotFoundError(f"Missing input workbook for {scenario_name}: {input_path}")

        excel_cells = read_sheet_cells(recalc_path, "Output", output_cells)
        process = next(
            (
                clean_text(excel_cells[address])
                for address in ("J61", "K61", "L61", "M61")
                if clean_text(excel_cells[address]) in PROCESS_COL
            ),
            "",
        )
        if not process:
            raise ValueError(f"Could not determine selected process for {scenario_name}")

        scenario = scenario_from_input_workbook(input_path, process)
        python_output = python_ported_output_summary_table(scenario).set_index(CommonColumns.METRIC)

        selected_column = PROCESS_COL[process]
        for metric, row in SELECTED_RECYCLING_METRICS.items():
            excel_value = clean_number(excel_cells[f"{selected_column}{row}"])
            python_value = clean_number(python_output.loc[metric, process])
            if excel_value is None or python_value is None:
                continue
            selected_rows.append(
                {
                    "scenario": scenario_name,
                    "process": process,
                    "metric": metric,
                    "excel": excel_value,
                    "python": python_value,
                    "delta": python_value - excel_value,
                    "abs_delta": abs(python_value - excel_value),
                }
            )

        for metric, route, address in OUTPUT_SUMMARY_CELLS:
            excel_value = clean_number(excel_cells[address])
            try:
                python_value = clean_number(python_output.loc[metric, route])
            except KeyError:
                python_value = None

            status = "compared" if excel_value is not None and python_value is not None else "skipped_non_numeric"
            delta = python_value - excel_value if status == "compared" else None
            full_rows.append(
                {
                    "scenario": scenario_name,
                    "process": process,
                    "metric": metric,
                    "route": route,
                    "cell": address,
                    "excel": excel_value,
                    "python": python_value,
                    "delta": delta,
                    "abs_delta": abs(delta) if delta is not None else None,
                    "status": status,
                }
            )

    selected = pd.DataFrame(selected_rows).sort_values("abs_delta", ascending=False)
    full = pd.DataFrame(full_rows).sort_values(
        ["status", "abs_delta"],
        ascending=[True, False],
        na_position="last",
    )
    return selected, full


def summarize(selected: pd.DataFrame, full: pd.DataFrame, args: argparse.Namespace) -> dict[str, object]:
    compared_full = full[full["status"] == "compared"]
    return {
        "scenarios": int(selected["scenario"].nunique()) if not selected.empty else 0,
        "selected_points": int(len(selected)),
        "selected_tolerance": args.selected_tolerance,
        "selected_over_tolerance": int((selected["abs_delta"] > args.selected_tolerance).sum()),
        "selected_max_abs_delta_by_metric": selected.groupby("metric")["abs_delta"].max().to_dict()
        if not selected.empty
        else {},
        "full_compared_points": int(len(compared_full)),
        "full_skipped_non_numeric": int((full["status"] == "skipped_non_numeric").sum()),
        "full_tolerance": args.full_tolerance,
        "full_over_tolerance": int((compared_full["abs_delta"] > args.full_tolerance).sum()),
        "full_max_abs_delta_by_metric": compared_full.groupby("metric")["abs_delta"].max().to_dict()
        if not compared_full.empty
        else {},
    }


def main() -> None:
    args = parse_args()
    if args.generate_inputs:
        generate_matrix_inputs(args)
    if args.recalculate:
        run_libreoffice_recalc(args)

    args.out_dir.mkdir(parents=True, exist_ok=True)
    selected, full = compare_workbooks(args.input_dir, args.recalc_dir)

    selected_path = args.out_dir / "selected_recycling_compare.csv"
    full_path = args.out_dir / "output_summary_compare.csv"
    summary_path = args.out_dir / "summary.json"
    selected.to_csv(selected_path, index=False)
    full.to_csv(full_path, index=False)

    summary = summarize(selected, full, args)
    summary.update(
        {
            "input_dir": str(args.input_dir),
            "recalc_dir": str(args.recalc_dir),
            "selected_csv": str(selected_path),
            "full_csv": str(full_path),
        }
    )
    summary_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")
    print(json.dumps(summary, indent=2))

    if args.strict and summary["selected_over_tolerance"]:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
