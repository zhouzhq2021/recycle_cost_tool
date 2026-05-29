from __future__ import annotations

from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from .schemas import CommonColumns, ManufacturingColumns
from .workbook import cell_value, load_everbatt_workbook


def number(value: Any, default: float = 0.0) -> float:
    if value is None:
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def text(value: Any, default: str = "") -> str:
    if value is None:
        return default
    return str(value).strip()


def yes(value: Any) -> bool:
    return text(value).casefold() == "yes"


def workbook_sheet(sheet: str, data_only: bool = True) -> Worksheet:
    return load_everbatt_workbook(data_only=data_only)[sheet]


def workbook_number(sheet: str, address: str, default: float = 0.0, data_only: bool = True) -> float:
    return number(cell_value(sheet, address, data_only=data_only), default)


def workbook_text(sheet: str, address: str, default: str = "", data_only: bool = True) -> str:
    return text(cell_value(sheet, address, data_only=data_only), default)


def workbook_range_text_values(sheet: str, cells: str) -> tuple[str, ...]:
    ws = workbook_sheet(sheet)
    values = []
    for row in ws[cells]:
        for cell in row:
            if cell.value is not None:
                values.append(str(cell.value))
    return tuple(values)


def regional_cost_factor(location: str, default: float = 1.0) -> float:
    geo = workbook_sheet("Geographic Par.")
    for column in range(3, 10):
        if geo.cell(30, column).value == location:
            return number(geo.cell(31, column).value, default) or default
    return default

# --- Manufacturing Parameters ---
MAN_PAR_SHEET = "Man Par."
MAN_REC_PAR_SHEET = "Man Rec Par."

MANUFACTURING_CHEMISTRY_HEADERS = ("Selected", "NMC(111)", "NMC(532)", "NMC(622)", "NMC(811)", "LCO", "NCA", "LMO", "LFP")
MANUFACTURING_RECYCLING_PROCESSES = ("Pyro", "Hydro", "Direct", "Custom")


def get_manufacturing_locations() -> dict[str, str]:
    return {
        "us": workbook_text("Input", "AD35"),
        "china": workbook_text("Input", "AD36"),
        "korea": workbook_text("Input", "AD37"),
        "europe": workbook_text("Input", "AD38"),
    }


def get_manufacturing_energy_factors(kind: str) -> dict[str, float | str]:
    if kind == "recycled":
        return {
            "total_energy": workbook_number(MAN_REC_PAR_SHEET, "AC40"),
            "electricity_share": workbook_number(MAN_REC_PAR_SHEET, "AC41"),
            "fuel_share": workbook_number(MAN_REC_PAR_SHEET, "AC42"),
            "mmbtu_to_mj": workbook_number(MAN_REC_PAR_SHEET, "AQ70"),
            "location": workbook_text(MAN_REC_PAR_SHEET, "AC10"),
        }
    return {
        "total_energy": workbook_number(MAN_PAR_SHEET, "C37"),
        "electricity_share": workbook_number(MAN_PAR_SHEET, "C38"),
        "fuel_share": workbook_number(MAN_PAR_SHEET, "C39"),
        "mmbtu_to_mj": workbook_number(MAN_PAR_SHEET, "H67"),
        "location": workbook_text(MAN_PAR_SHEET, "C10"),
    }


def get_manufacturing_cathode_chemistry() -> str:
    return workbook_text(MAN_REC_PAR_SHEET, "AC9")


def get_manufacturing_recycled_share() -> float:
    return workbook_number(MAN_REC_PAR_SHEET, "AC11")


def get_manufacturing_cathode_conversion_factor() -> float:
    return workbook_number(MAN_REC_PAR_SHEET, "AQ69")


def get_manufacturing_general_inputs_addresses(kind: str) -> tuple[tuple[str, str, str], ...]:
    if kind == "virgin":
        return (
            ("throughput", "tonne/yr", "C8"),
            ("cathode_chemistry", "", "C9"),
            ("geographic_location", "", "C10"),
        )
    return (
        ("throughput", "tonne/yr", "AC8"),
        ("cathode_chemistry", "", "AC9"),
        ("geographic_location", "", "AC10"),
        ("recycled_content_selected_chemistry", "", "AC11"),
        ("recycled_content_other_materials_1", "", "AC12"),
        ("recycled_content_other_materials_2", "", "AC13"),
    )


def get_manufacturing_wide_table_specs(kind: str, table_name: str) -> dict:
    if kind == "virgin":
        specs = {
            "cell_size": (2, range(3, 12), range(15, 17), CommonColumns.ITEM),
            "cell_material_composition": (2, range(3, 12), range(20, 34), CommonColumns.MATERIAL),
            "cell_energy_consumption": (2, range(3, 12), range(37, 40), CommonColumns.ITEM),
            "module_component_masses": (28, range(29, 38), range(19, 25), CommonColumns.COMPONENT),
            "pack_component_masses": (28, range(29, 38), range(28, 41), CommonColumns.COMPONENT),
            "pack_energy_consumption": (28, range(29, 38), range(82, 84), CommonColumns.ITEM),
        }
    else:
        specs = {
            "cell_size": (28, range(29, 38), range(18, 20), CommonColumns.ITEM),
            "cell_material_composition": (28, range(29, 38), range(23, 37), CommonColumns.MATERIAL),
            "cell_energy_consumption": (28, range(29, 38), range(40, 43), CommonColumns.ITEM),
        }
    return specs.get(table_name)


def get_manufacturing_yields_specs(kind: str) -> dict:
    if kind == "virgin":
        return {
            "label_col": 2,
            "value_cols": (3, 4, 5),
            "rows": (44, 47, 48, 49, 50, 51, 52),
        }
    return {
        "label_col": 28,
        "value_cols": (29, 30, 31),
        "rows": (47, 50, 51, 52, 53, 54, 55),
    }


def get_manufacturing_solvent_use_specs(kind: str) -> dict:
    if kind == "virgin":
        return {
            "label_col": 2,
            "solvent_col": 3,
            "ratio_col": 4,
            "recovery_col": 7,
            "rows": (56, 57),
        }
    return {
        "label_col": 28,
        "solvent_col": 29,
        "ratio_col": 30,
        "recovery_col": 33,
        "rows": (59, 60),
    }


def get_manufacturing_material_inputs_specs(kind: str) -> dict:
    if kind == "virgin":
        return {"header_row": 61, "value_row": 62, "first_col": 3, "last_col": 17}
    return {"header_row": 64, "value_row": 65, "first_col": 29, "last_col": 43}


def get_manufacturing_environment_summary_rows(kind: str) -> range:
    if kind == "virgin":
        return range(67, 87)
    return range(70, 90)


def get_manufacturing_material_cost_specs(kind: str) -> dict:
    if kind == "virgin":
        return {"header_row": 90, "value_row": 91, "first_col": 3, "last_col": 14}
    return {"header_row": 93, "value_row": 94, "first_col": 29, "last_col": 42}


def get_manufacturing_cost_summary_rows(kind: str) -> range:
    if kind == "virgin":
        return range(105, 120)
    return range(109, 124)


def get_manufacturing_recycled_cathode_material_environment_blocks() -> dict:
    return {
        "Pyro": (7, range(9, 29), range(3, 9)),
        "Hydro": (31, range(33, 53), range(3, 9)),
        "Direct": (55, range(57, 77), range(3, 11)),
        "Custom": (79, range(81, 101), range(3, 9)),
    }


def get_manufacturing_recycled_cathode_material_costs_blocks() -> dict:
    return {
        "Pyro": (103, 104, range(3, 9)),
        "Hydro": (107, 108, range(3, 9)),
        "Direct": (111, 112, range(3, 11)),
        "Custom": (115, 116, range(3, 9)),
        "Virgin": (119, 120, range(3, 11)),
    }


def get_pack_basic_inputs_addresses() -> tuple[tuple[str, str, str], ...]:
    return (
        ("throughput", "tonne battery pack/yr", "AC8"),
        ("geographic_location", "", "AC9"),
    )


# --- Cathode Parameters ---
CATHODE_PRODUCTION_SHEET = "Cath. Prod. Par."

CATHODE_PRODUCTION_BLOCKS = {
    "NMC(111)": 27,
    "NMC(532)": 53,
    "NMC(622)": 79,
    "NMC(811)": 105,
    "NCA": 131,
    "LCO": 157,
}

CATHODE_CHEMISTRY_COLUMNS = {
    "LCO": 3,
    "NMC(111)": 4,
    "NMC(532)": 5,
    "NMC(622)": 6,
    "NMC(811)": 7,
    "NCA": 8,
    "LMO": 9,
    "LFP": 10,
}

CATHODE_PROCESS_COLUMNS = {
    "Pyro": 3,
    "Hydro": 4,
    "Custom": 5,
}

CATHODE_PRICE_COLUMNS = {
    "selected": 3,
    "default": 4,
    "user_defined": 5,
}

CATHODE_REQUIRED_PRECURSOR_COLUMNS = {
    "selected": 1,
    "default_greet": 2,
    "default_builder": 3,
    "user_defined": 4,
}

CATHODE_SPLIT_OFFSETS = {
    "Pyro": (1, 2, 3),
    "Hydro": (4, 5, 6),
    "Custom": (7, 8, 9),
}

CATHODE_ENVIRONMENT_COLUMNS = {
    "material_pyro": 1,
    "material_hydro": 2,
    "material_custom": 3,
    "energy_input": 4,
    "process": 5,
    "total_pyro": 6,
    "total_hydro": 7,
    "total_custom": 8,
}

CATHODE_COST_PER_LINE_COLUMNS = {
    "Pyro": 1,
    "Hydro": 2,
    "Custom": 3,
    "Virgin": 4,
}

CATHODE_DETAILED_COST_ROWS = {
    483: "Total capital investment",
    485: "Manufacturing cost",
    486: "Direct product costs",
    488: "Raw materials",
    490: "Operating labor",
    492: "Direct supervisory and clerical labor",
    494: "Utilities",
    496: "Maintenance and repairs",
    498: "Operating supplies",
    500: "Laboratory charges",
    502: "Patents and royalties",
    503: "Fixed charges",
    505: "Depreciation",
    507: "Local taxes",
    509: "Insurance",
    511: "Rent",
    513: "Financing",
    515: "Plant overhead costs",
    516: "General expenses",
    518: "Administrative costs",
    520: "Distribution and selling costs",
    522: "R&D costs",
    524: "Total product cost",
    526: "Profit",
    527: "Total product cost w/profit",
    528: "Total product cost to recipient",
}


def get_cathode_general_inputs() -> dict[str, str | float]:
    return {
        "throughput": workbook_number(CATHODE_PRODUCTION_SHEET, "D8"),
        "cathode_chemistry": workbook_text(CATHODE_PRODUCTION_SHEET, "D9"),
        "geographic_location": workbook_text(CATHODE_PRODUCTION_SHEET, "D10"),
    }


def get_cathode_tonnes_per_gwh_factors() -> dict[str, float]:
    return {
        "mass_per_kwh": workbook_number("Man Rec Par.", "AC65"),
        "active_share": workbook_number("Man Rec Par.", "AC19"),
    }


def get_cathode_precursor_rows(start: int = 14, end: int = 19) -> list[tuple[str, dict]]:
    ws = workbook_sheet(CATHODE_PRODUCTION_SHEET)
    records: list[tuple[str, dict]] = []
    for row in range(start, end + 1):
        precursor = ws.cell(row, 2).value
        if precursor is None:
            continue
        record: dict = {}
        for col in range(3, 8):
            record[col] = number(ws.cell(row, col).value)
        records.append((str(precursor), record))
    return records


def get_cathode_throughput_address() -> str:
    return "D8"


def get_cathode_general_default() -> dict[str, object]:
    return get_cathode_general_inputs()


def get_cathode_material_energy_demand_data() -> list[dict]:
    ws = workbook_sheet(CATHODE_PRODUCTION_SHEET)
    records = []
    for row in range(27, 49):
        item = ws.cell(row, 2).value
        record = {CommonColumns.ITEM: item, "row": row}
        for chemistry, col in CATHODE_CHEMISTRY_COLUMNS.items():
            record[chemistry] = number(ws.cell(row, col).value)
        records.append(record)
    return records


def get_cathode_chemical_prices_data() -> list[dict]:
    ws = workbook_sheet(CATHODE_PRODUCTION_SHEET)
    records = []
    for row in range(52, 73):
        chemical = ws.cell(row, 2).value
        record = {"chemical": chemical, "row": row}
        for label, col in CATHODE_PRICE_COLUMNS.items():
            record[label] = number(ws.cell(row, col).value)
        records.append(record)
    return records


def get_cathode_utility_prices_data() -> list[dict]:
    ws = workbook_sheet(CATHODE_PRODUCTION_SHEET)
    records = []
    for row in range(76, 79):
        utility = ws.cell(row, 2).value
        record = {"utility": utility, "row": row}
        for label, col in CATHODE_PRICE_COLUMNS.items():
            record[label] = number(ws.cell(row, col).value)
        records.append(record)
    return records


def get_cathode_material_conversion_costs_data() -> list[dict]:
    ws = workbook_sheet(CATHODE_PRODUCTION_SHEET)
    records = []
    for col in range(3, 7):
        precursor = ws.cell(81, col).value
        records.append(
            {
                CommonColumns.MATERIAL: precursor,
                "cost_per_kg_precursor": number(ws.cell(82, col).value),
            }
        )
    return records


def get_cathode_direct_regeneration_data() -> dict:
    mat_conv = workbook_sheet("Mat. Conv Par.")
    direct_quantities = {
        str(mat_conv.cell(row, 2).value): number(mat_conv.cell(row, 5).value)
        for row in range(14, 22)
        if mat_conv.cell(row, 2).value is not None
    }
    recovered_prices = {
        str(mat_conv.cell(52, col).value): number(mat_conv.cell(53, col).value)
        for col in range(7, 15)
        if mat_conv.cell(52, col).value is not None
    }
    return {
        "direct_quantities": direct_quantities,
        "recovered_prices": recovered_prices,
        "raw_data": mat_conv,
    }


# --- Preprocessing Parameters ---
def get_preproc_throughputs() -> dict[str, float]:
    return {
        "generic": workbook_number("Preproc. Par.", "AD14"),
        "specific": workbook_number("Preproc. Par.", "BD15"),
        "specific_old": workbook_number("Preproc. Par.", "BD14"),
    }


def get_preproc_energy_demand_mmbtu() -> dict[str, float]:
    return {
        "generic": workbook_number("Preproc. Par.", "AC195"),
        "specific": workbook_number("Preproc. Par.", "BC195"),
    }


def get_preproc_ghg_factors() -> dict[str, float]:
    return {
        "generic": workbook_number("Preproc. Par.", "AK184"),
        "specific": workbook_number("Preproc. Par.", "BL184"),
    }


def get_preproc_default_value() -> float:
    return workbook_number("Preproc. Par.", "C89")


# --- UI Defaults ---
def get_input_selected_chemistry() -> str:
    return workbook_text("Input", "E38")


# --- Reporting Parameters ---
REPORTING_MANUFACTURING_OUTPUT_SUMMARY_SPECS = [
    ("Cell manufacturing cost", CommonColumns.COST, "per kWh battery produced", "Cost per kWh batttery produced"),
    ("Cell manufacturing total energy", "Energy", "MJ per kWh battery produced", "Total Energy"),
    ("Cell manufacturing water", "Water", "gal per kWh battery produced", "Water use in gallon"),
    ("Cell manufacturing NOx", "Emissions", "g per kWh battery produced", "NOx"),
    ("Cell manufacturing PM10", "Emissions", "g per kWh battery produced", "PM10"),
    ("Cell manufacturing SOx", "Emissions", "g per kWh battery produced", "SOx"),
    ("Cell manufacturing GHGs", "Emissions", "g CO2e per kWh battery produced", "GHGs"),
]

REPORTING_RECYCLING_OUTPUT_SUMMARY_SPECS = [
    ("Recycling cost", CommonColumns.COST, "per kg feedstock processed", "Cost per kg feedstock processed"),
    ("Recycling total energy", "Energy", "MJ per kg feedstock processed", "Total Energy"),
    ("Recycling water", "Water", "gal per kg feedstock processed", "Water use in gallon"),
    ("Recycling revenue", CommonColumns.REVENUE, "per kg feedstock processed", "Revenue per kg feedstock processed"),
    ("Recycling GHGs", "Emissions", "g CO2e per kg feedstock processed", "GHGs"),
]


def get_reporting_recycle_columns() -> dict:
    return {
        "Pyro": {"mode_col": 55, "throughput_col": 56, "cost_row": 197, "env_col": 58, CommonColumns.REVENUE + "_col": 62},
        "Hydro": {"mode_col": 81, "throughput_col": 82, "cost_row": 197, "env_col": 84, CommonColumns.REVENUE + "_col": 90},
        "Direct": {"mode_col": 107, "throughput_col": 108, "cost_row": 197, "env_col": 110, CommonColumns.REVENUE + "_col": 116},
        "Custom": {"mode_col": 133, "throughput_col": 134, "cost_row": 199, "env_col": 136, CommonColumns.REVENUE + "_col": 142},
    }


def get_reporting_recycling_process_specs() -> dict:
    return {
        "Pyro": {"mode": "BC8", "throughput": "BD25", "value_col": 59, "capital_col": 60, "feed": "C89"},
        "Hydro": {"mode": "CC8", "throughput": "CD25", "value_col": 85, "capital_col": 86, "feed": "D89"},
        "Direct": {"mode": "DC8", "throughput": "DD25", "value_col": 111, "capital_col": 112, "feed": "E89"},
        "Custom": {"mode": "EC8", "throughput": "ED26", "value_col": 137, "capital_col": 138, "feed": "F89"},
    }


def get_reporting_recycling_item_specs() -> dict:
    return {
        "Materials": (444, 443, 478, "value_col"),
        "Utilities": (445, 444, 479, "value_col"),
        "Other variable costs": (447, 446, 481, "value_col"),
        "Labor": (450, 449, 484, "value_col"),
        "Maintenance": (454, 453, 488, "value_col"),
        "Plant overhead": (457, 456, 491, "value_col"),
        "Other fixed costs": (458, 457, 492, "value_col"),
        "Annualized capital cost": (467, 466, 501, "capital_col"),
    }
