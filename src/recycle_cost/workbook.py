from __future__ import annotations

import re
import tempfile
import zipfile
from functools import lru_cache
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


ROOT = Path(__file__).resolve().parents[2]
WORKBOOK_PATH = ROOT / "EverBatt 2023.xlsm"
ANALYSIS_PATH = ROOT / "docs" / "analysis" / "everbatt_analysis.json"


def _clamp_font_family(match: re.Match[bytes]) -> bytes:
    value = int(match.group(1))
    if value <= 14:
        return match.group(0)
    return b'<family val="14"'


def make_openpyxl_readable_copy(source: Path = WORKBOOK_PATH) -> Path:
    """Create a temporary xlsm copy with XML style values normalized for openpyxl."""
    tmp = tempfile.NamedTemporaryFile(
        prefix="everbatt_readable_",
        suffix=".xlsm",
        delete=False,
    )
    tmp_path = Path(tmp.name)
    tmp.close()

    with zipfile.ZipFile(source) as archive:
        with zipfile.ZipFile(tmp_path, "w", compression=zipfile.ZIP_DEFLATED) as out:
            for info in archive.infolist():
                data = archive.read(info.filename)
                if info.filename.endswith(".xml"):
                    data = re.sub(rb'<family val="([0-9]+)"', _clamp_font_family, data)
                out.writestr(info, data)

    return tmp_path


@lru_cache(maxsize=2)
def load_everbatt_workbook(data_only: bool = True) -> Workbook:
    readable = make_openpyxl_readable_copy()
    return load_workbook(readable, data_only=data_only, keep_vba=True)


def cell_value(sheet: str, address: str, data_only: bool = True) -> Any:
    wb = load_everbatt_workbook(data_only=data_only)
    return wb[sheet][address].value


def worksheet_to_frame(
    sheet: str,
    min_row: int = 1,
    max_row: int | None = None,
    min_col: int = 1,
    max_col: int | None = None,
    data_only: bool = True,
) -> pd.DataFrame:
    ws = load_everbatt_workbook(data_only=data_only)[sheet]
    return worksheet_range_to_frame(ws, min_row, max_row, min_col, max_col)


def worksheet_range_to_frame(
    ws: Worksheet,
    min_row: int = 1,
    max_row: int | None = None,
    min_col: int = 1,
    max_col: int | None = None,
) -> pd.DataFrame:
    max_row = max_row or ws.max_row
    max_col = max_col or ws.max_column
    rows = []
    for row in ws.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        values_only=True,
    ):
        rows.append(list(row))
    return pd.DataFrame(rows)

