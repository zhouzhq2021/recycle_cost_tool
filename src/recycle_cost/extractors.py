from __future__ import annotations

from dataclasses import dataclass
from typing import Iterable

import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet

from .workbook import load_everbatt_workbook


REGIONS = ["U.S.", "California", "China", "Korea"]
POLLUTANTS = ["VOC", "CO", "NOx", "PM10", "PM2.5", "SOx", "BC", "OC", "CH4", "N2O", "CO2"]


@dataclass(frozen=True)
class TableBlock:
    name: str
    sheet: str
    start_row: int
    end_row: int
    label_col: int
    selected_col: int | None = None
    default_col: int | None = None
    unit: str | None = None


def _ws(sheet: str, data_only: bool = True) -> Worksheet:
    return load_everbatt_workbook(data_only=data_only)[sheet]


def _nonempty(value) -> bool:
    return value is not None and value != ""


def extract_material_prices() -> pd.DataFrame:
    ws = _ws("Materials", data_only=True)
    blocks = [
        TableBlock("Metals", "Materials", 4, 6, 1, selected_col=2, default_col=3, unit="$/kg"),
        TableBlock("Recycling chemicals", "Materials", 10, 31, 1, selected_col=2, default_col=3, unit="$/kg"),
        TableBlock("Material conversion chemicals", "Materials", 35, 80, 1, selected_col=2, default_col=3, unit="$/kg"),
        TableBlock("Recovered materials", "Materials", 98, 117, 1, selected_col=2, default_col=3, unit="$/kg"),
    ]
    rows = []
    for block in blocks:
        for row in range(block.start_row, block.end_row + 1):
            name = ws.cell(row, block.label_col).value
            if not _nonempty(name):
                continue
            selected = ws.cell(row, block.selected_col).value if block.selected_col else None
            default = ws.cell(row, block.default_col).value if block.default_col else None
            rows.append(
                {
                    "group": block.name,
                    "material": name,
                    "selected": selected,
                    "default": default,
                    "unit": block.unit,
                    "cell": f"Materials!A{row}",
                }
            )
    return pd.DataFrame(rows)


def extract_geographic_parameters() -> pd.DataFrame:
    ws = _ws("Geographic Par.", data_only=True)
    sections = [
        ("Battery manufacturing", 8, 10),
        ("Rail, barge, ocean tanker transport", 16, 19),
        ("Truck transport", 23, 26),
        ("Battery recycling", 31, 38),
        ("Cathode production", 42, 45),
        ("Battery manufacturing with recycled materials", 72, 74),
    ]
    rows = []
    for section, start, end in sections:
        for row in range(start, end + 1):
            label = ws.cell(row, 2).value
            if not _nonempty(label):
                continue
            values = [ws.cell(row, col).value for col in range(4, 8)]
            if all(value is None for value in values):
                values = [ws.cell(row, col).value for col in range(3, 7)]
            for region, value in zip(REGIONS, values, strict=True):
                rows.append(
                    {
                        "section": section,
                        "parameter": label,
                        "region": region,
                        "value": value,
                        "cell": f"Geographic Par.!{row}",
                    }
                )
    return pd.DataFrame(rows)


def extract_batpac_material_costs() -> pd.DataFrame:
    ws = _ws("BatPaC IO", data_only=True)
    rows = []
    for col in range(2, ws.max_column + 1):
        material = ws.cell(4, col).value
        cost = ws.cell(5, col).value
        if _nonempty(material):
            rows.append(
                {
                    "material": material,
                    "cost_per_kg": cost,
                    "cell": f"BatPaC IO!{ws.cell(5, col).coordinate}",
                }
            )
    return pd.DataFrame(rows)


def extract_greet_combustion_factors() -> pd.DataFrame:
    ws = _ws("GREET IO", data_only=True)
    fuels = []
    current_fuel = None
    for col in range(2, 38):
        maybe_fuel = ws.cell(2, col).value
        if _nonempty(maybe_fuel):
            current_fuel = maybe_fuel
        technology = ws.cell(3, col).value
        if _nonempty(technology) and current_fuel:
            fuels.append((col, current_fuel, technology))

    rows = []
    for row in range(4, 15):
        pollutant = ws.cell(row, 1).value
        if pollutant not in POLLUTANTS:
            continue
        for col, fuel, technology in fuels:
            rows.append(
                {
                    "fuel": fuel,
                    "technology": technology,
                    "pollutant": pollutant,
                    "value": ws.cell(row, col).value,
                    "unit": "g/mmBtu fuel burned",
                    "cell": f"GREET IO!{ws.cell(row, col).coordinate}",
                }
            )
    return pd.DataFrame(rows)


def available_reference_tables() -> dict[str, pd.DataFrame]:
    return {
        "Material prices": extract_material_prices(),
        "Geographic parameters": extract_geographic_parameters(),
        "BatPaC material costs": extract_batpac_material_costs(),
        "GREET combustion factors": extract_greet_combustion_factors(),
    }

